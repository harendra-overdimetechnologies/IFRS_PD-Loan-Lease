import pandas as pd
import os
import glob
import logging
import argparse
import openpyxl
from pyxlsb import open_workbook
from datetime import datetime


class LoggerSetup:
    """
    Class to handle logging configuration and setup
    """
    def __init__(self, log_folder, log_prefix="PD_Loan"):
        self.log_folder = log_folder
        self.log_prefix = log_prefix
        self.logger = None
        self._setup_logger()
    
    def _setup_logger(self):
        """Setup and configure logger"""
        os.makedirs(self.log_folder, exist_ok=True)
        
        log_filename = f"{self.log_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        log_path = os.path.join(self.log_folder, log_filename)
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_path),
                logging.StreamHandler()
            ]
        )
        
        self.logger = logging.getLogger(__name__)
    
    def get_logger(self):
        """Return the configured logger"""
        return self.logger


class FileHandler:
    """
    Class to handle file operations like searching and reading Excel files
    """
    def __init__(self, logger):
        self.logger = logger
    
    def find_files_by_keyword(self, base_folder, keyword):
        """
        Find files in the base folder that contain the keyword in their name
        
        Args:
            base_folder (str): Base folder to search in
            keyword (str): Keyword to search for in file names
            
        Returns:
            list: List of matching file paths
        """
        self.logger.info(f"Searching for files with keyword '{keyword}' in {base_folder}")
        
        search_pattern = os.path.join(base_folder, "**", "*" + keyword + "*.xlsx*")
        matching_files = glob.glob(search_pattern, recursive=True)
        
        # Exclude Excel temp lock files (start with ~$)
        matching_files = [f for f in matching_files if not os.path.basename(f).startswith('~$')]
        
        self.logger.info(f"Found {len(matching_files)} matching file(s)")
        for file in matching_files:
            self.logger.info(f"  - {file}")
        
        return matching_files
    
    def read_excel_sheet(self, file_path, sheet_name, header_row=None, column_names=None, skiprows=None, nrows=None):
        """
        Read a specific sheet from an Excel file
        
        Args:
            file_path (str): Path to the Excel file
            sheet_name (str): Name of the sheet to read
            header_row (int): Row number to use as header (0-indexed)
            column_names (list): List of column names to set
            skiprows (int or list): Rows to skip at the beginning
            nrows (int): Number of rows to read
            
        Returns:
            pd.DataFrame: DataFrame containing the sheet data
        """
        self.logger.info(f"Reading '{sheet_name}' sheet from: {file_path}")
        
        try:
            df = pd.read_excel(
                file_path, 
                sheet_name=sheet_name, 
                header=header_row,
                skiprows=skiprows,
                nrows=nrows
            )
            
            # Set column names if provided
            if column_names:
                df.columns = column_names
                self.logger.info(f"Set custom column names: {column_names}")
            
            self.logger.info(f"Successfully read '{sheet_name}' sheet")
            self.logger.info(f"DataFrame shape: {df.shape}")
            self.logger.info(f"Columns: {list(df.columns)}")
            self.logger.info(f"Data range: Excel rows {skiprows + 1 if skiprows else 0} to {(skiprows or 0) + (nrows or len(df))}")
            
            return df
        except Exception as e:
            self.logger.error(f"Error reading file: {str(e)}")
            raise


class HistoricPDPivotExtractor:
    """Step 2: Extract per-category values equivalent to 03.PD_Pivot!M5:M10."""

    def __init__(self, input_folder, logger):
        self.input_folder = input_folder
        self.logger = logger
        self.target_file = None

    def find_historic_file(self, keyword="Historic PD Calculation"):
        pattern = os.path.join(self.input_folder, "**", f"*{keyword}*.xlsb")
        matches = glob.glob(pattern, recursive=True)
        matches = [m for m in matches if not os.path.basename(m).startswith("~$")]

        self.logger.info("Searching file with keyword '%s' in %s", keyword, self.input_folder)
        self.logger.info("Found %s matching file(s)", len(matches))
        for m in matches:
            self.logger.info("  - %s", m)

        if not matches:
            raise FileNotFoundError(f"No .xlsb file found for keyword: {keyword}")

        self.target_file = matches[0]
        return self.target_file

    def _read_working_dataframe(self):
        """Read 02.Working and keep fields needed for category-wise M5:M10 logic."""
        if not self.target_file:
            raise ValueError("Target file not set. Run find_historic_file() first.")

        rows = []
        with open_workbook(self.target_file) as wb:
            with wb.get_sheet("02.Working") as sh:
                for r_idx, row in enumerate(sh.rows(), start=1):
                    vals = [c.v for c in row]

                    # Header row in 02.Working is row 2
                    if r_idx <= 2:
                        continue

                    # Columns: A=contract, B=PD_CATEGORY, Q=WORST, R=FIRST
                    contract = vals[0] if len(vals) > 0 else None
                    category = vals[1] if len(vals) > 1 else None
                    worst = vals[16] if len(vals) > 16 else None
                    first = vals[17] if len(vals) > 17 else None

                    if contract is None and category is None:
                        continue

                    rows.append(
                        {
                            "CONTRACT": contract,
                            "PD_CATEGORY": category,
                            "WORST": worst,
                            "FIRST": first,
                        }
                    )

        df = pd.DataFrame(rows)
        df = df[df["PD_CATEGORY"].notna()].copy()

        # Normalize bucket fields to integers 1..5 where possible
        df["FIRST"] = pd.to_numeric(df["FIRST"], errors="coerce")
        df["WORST"] = pd.to_numeric(df["WORST"], errors="coerce")
        df = df[df["FIRST"].isin([1, 2, 3, 4, 5])]
        df = df[df["WORST"].isin([1, 2, 3, 4, 5])]
        df["FIRST"] = df["FIRST"].astype(int)
        df["WORST"] = df["WORST"].astype(int)

        self.logger.info("Loaded 02.Working records: %s", len(df))
        return df

    def _extract_from_pivot_formula_values(self):
        """
        Extract M5:M10 values directly from 03.PD_Pivot formula results
        by switching PD_CATEGORY in the pivot filter.

        This returns the exact cached/recalculated values that Excel shows,
        including WAPD in row 10.
        """
        if not self.target_file:
            raise ValueError("Target file not set. Run find_historic_file() first.")

        try:
            import win32com.client as win32
        except Exception as e:
            self.logger.warning("win32com not available, fallback to computed values: %s", e)
            return None

        excel = None
        wb = None
        result = {}

        try:
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            wb = excel.Workbooks.Open(self.target_file)
            ws = wb.Worksheets("03.PD_Pivot")

            pivot = None
            for i in range(1, ws.PivotTables().Count + 1):
                pt = ws.PivotTables(i)
                try:
                    _ = pt.PivotFields("PD_CATEGORY")
                    pivot = pt
                    break
                except Exception:
                    continue

            if pivot is None:
                self.logger.warning("No pivot table with PD_CATEGORY found in 03.PD_Pivot")
                return None

            pf = pivot.PivotFields("PD_CATEGORY")
            categories = [pf.PivotItems(i).Name for i in range(1, pf.PivotItems().Count + 1)]
            categories = [str(c).strip() for c in categories if str(c).strip() and str(c).strip().lower() != "(blank)"]

            self.logger.info("Formula extraction categories from pivot: %s", categories)

            for cat in categories:
                # Ensure target category is visible first
                for i in range(1, pf.PivotItems().Count + 1):
                    item = pf.PivotItems(i)
                    if str(item.Name).strip() == cat:
                        item.Visible = True
                        break

                # Hide all other categories
                for i in range(1, pf.PivotItems().Count + 1):
                    item = pf.PivotItems(i)
                    if str(item.Name).strip() != cat:
                        item.Visible = False

                pivot.RefreshTable()
                excel.Calculate()

                m_vals = ws.Range("M5:M10").Value
                k_vals = ws.Range("K5:K10").Value

                out_rows = []
                dc_buckets = [1, 2, 3, 4, 5, "WAPD"]
                for idx, dc in enumerate(dc_buckets):
                    m_raw = m_vals[idx][0] if m_vals and m_vals[idx] else None
                    k_raw = k_vals[idx][0] if k_vals and k_vals[idx] else None

                    m_value = 0 if m_raw is None else float(m_raw)
                    contracts = 0 if k_raw is None else int(float(k_raw))

                    out_rows.append(
                        {
                            "PD_CATEGORY": cat,
                            "excel_row": 5 + idx,
                            "DC_BUCKET": dc,
                            "M_value": m_value,
                            "contracts_in_bucket": contracts,
                        }
                    )

                result[cat] = pd.DataFrame(out_rows)

            return result

        except Exception as e:
            self.logger.warning("Formula-based pivot extraction failed, fallback to computed values: %s", e)
            return None
        finally:
            try:
                if wb is not None:
                    wb.Close(False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass

    def _compute_category_dataframe(self, df_working, category):
        """
        Build dataframe equivalent to 03.PD_Pivot M5:M10 for one category.

        M5..M9 are interpreted as FINAL HISTORIC PD for DC Bucket 1..5:
            P(WORST=5 | FIRST=bucket)

        M10 is interpreted as WAPD (overall for the selected category):
            P(WORST=5)
        """
        cat_df = df_working[df_working["PD_CATEGORY"] == category]

        out_rows = []
        for bucket in [1, 2, 3, 4, 5]:
            bucket_df = cat_df[cat_df["FIRST"] == bucket]
            total = len(bucket_df)
            default_cnt = len(bucket_df[bucket_df["WORST"] == 5])
            m_value = (default_cnt / total) if total > 0 else 0

            out_rows.append(
                {
                    "PD_CATEGORY": category,
                    "excel_row": bucket + 4,
                    "DC_BUCKET": bucket,
                    "M_value": m_value,
                    "contracts_in_bucket": total,
                }
            )

        # Row 10 in pivot M column: WAPD for non-default opening buckets (FIRST 1..4)
        # This aligns with workbook logic where starting default bucket (FIRST=5)
        # is excluded from WAPD computation.
        non_default_opening_df = cat_df[cat_df["FIRST"].isin([1, 2, 3, 4])]
        total_non_default_opening = len(non_default_opening_df)
        default_from_non_default_opening = len(non_default_opening_df[non_default_opening_df["WORST"] == 5])
        m10_value = (
            default_from_non_default_opening / total_non_default_opening
            if total_non_default_opening > 0
            else 0
        )
        out_rows.append(
            {
                "PD_CATEGORY": category,
                "excel_row": 10,
                "DC_BUCKET": "WAPD",
                "M_value": m10_value,
                "contracts_in_bucket": total_non_default_opening,
            }
        )

        return pd.DataFrame(out_rows)

    def extract_all_categories(self):
        """Return dict: {category: dataframe for row 5..10 M column values}."""
        # Primary path: exact formula/cached values from 03.PD_Pivot
        pivot_result = self._extract_from_pivot_formula_values()
        if pivot_result:
            for cat, df_cat in pivot_result.items():
                self.logger.info("Category '%s' -> M(row 5..10) values (formula-derived):", cat)
                self.logger.info("\n%s", df_cat[["excel_row", "DC_BUCKET", "M_value"]].to_string(index=False))
            return pivot_result

        # Fallback path: compute from 02.Working
        df_working = self._read_working_dataframe()
        categories = sorted(df_working["PD_CATEGORY"].dropna().astype(str).unique().tolist())

        self.logger.info("Identified categories from 'Select PD Category' source: %s", categories)

        result = {}
        for cat in categories:
            df_cat = self._compute_category_dataframe(df_working, cat)
            result[cat] = df_cat

            self.logger.info("Category '%s' -> M(row 5..10) values:", cat)
            self.logger.info("\n%s", df_cat[["excel_row", "DC_BUCKET", "M_value"]].to_string(index=False))

        return result


class PDLoanAnalyzer:
    """
    Main class to orchestrate PD Loan analysis
    """
    def __init__(self, input_folder, log_folder):
        self.input_folder = input_folder
        self.log_folder = log_folder
        
        # Initialize logger
        logger_setup = LoggerSetup(log_folder)
        self.logger = logger_setup.get_logger()
        
        # Initialize file handler
        self.file_handler = FileHandler(self.logger)
        
        # Data storage
        self.df_economic_factors = None
        self.target_file = None
        self.historic_pd_frames = None
    
    def find_pd_loan_file(self, keyword="PD Loan"):
        """
        Find PD Loan file by keyword
        
        Args:
            keyword (str): Keyword to search for (default: "PD Loan")
            
        Returns:
            str: Path to the target file or None
        """
        matching_files = self.file_handler.find_files_by_keyword(self.input_folder, keyword)
        
        if not matching_files:
            self.logger.warning("No files found with the specified keyword")
            return None
        
        # Use the first matching file
        self.target_file = matching_files[0]
        self.logger.info(f"Using file: {self.target_file}")
        
        return self.target_file
    
    def load_economic_factors(self):
        """
        Load Economic Factors sheet from the target file
        Read data from Excel rows 3 to 15 (13 rows total)
        
        Returns:
            pd.DataFrame: DataFrame containing economic factors data
        """
        if not self.target_file:
            raise ValueError("No target file specified. Run find_pd_loan_file() first.")
        
        # Define proper column names (5 columns: Year and 4 economic factors)
        column_names = ['Year', 'GDP', 'CPI', 'Rf', 'Unemp']
        
        # Read data from Excel rows 3 to 15
        # skiprows=2 skips Excel rows 1-2 (header rows)
        # nrows=13 reads 13 rows (Excel rows 3 to 15)
        self.df_economic_factors = self.file_handler.read_excel_sheet(
            self.target_file, 
            'Economic Factors',
            header_row=None,
            column_names=column_names,
            skiprows=2,
            nrows=13
        )
        
        # Replace NaN values with 0
        self.df_economic_factors = self.df_economic_factors.fillna(0)
        
        self.logger.info(f"Loaded data from Excel rows 3 to 15")
        
        return self.df_economic_factors
    
    def parse_running_date(self, date_string):
        """
        Parse running date from string format (e.g., '9.03.2025')
        
        Args:
            date_string (str): Date in format DD.MM.YYYY
            
        Returns:
            datetime: Parsed date object
        """
        try:
            running_date = datetime.strptime(date_string, '%d.%m.%Y')
            self.logger.info(f"Running date: {running_date.strftime('%B %d, %Y')}")
            return running_date
        except ValueError as e:
            self.logger.error(f"Invalid date format. Expected DD.MM.YYYY: {e}")
            raise
    
    def update_with_shifting(self, year, gdp, cpi, rf, unemp):
        """
        Update dataframe with shifting method (for September):
          1. Drop the first row (oldest year)
          2. Shift all remaining rows up by one position
                    3. Append a new row at the end: Year = last_year + 1,
                         GDP/CPI/Rf/Unemp copied from row 11 (fallback: last row)
          4. Update the row matching the user-entered year with the user-entered values
        """
        self.logger.info("=" * 80)
        self.logger.info("SEPTEMBER UPDATE - Using Shifting Method")
        self.logger.info("=" * 80)

        # Step 1 & 2: Remove first row and shift up
        dropped_year = int(self.df_economic_factors.iloc[0]['Year'])
        self.df_economic_factors = self.df_economic_factors.iloc[1:].reset_index(drop=True)
        self.logger.info(f"Dropped first row (Year={dropped_year}) and shifted all rows up by one")

        # Step 3: Append new row at end with Year = last_year + 1,
        # using GDP/CPI/Rf/Unemp from row 11 (1-based) after shift.
        last_year = int(self.df_economic_factors.iloc[-1]['Year'])
        new_year = last_year + 1

        source_pos = 10  # Row 11 in 1-based counting
        if len(self.df_economic_factors) > source_pos:
            source_row = self.df_economic_factors.iloc[source_pos]
            source_row_no = source_pos + 1
        else:
            source_row = self.df_economic_factors.iloc[-1]
            source_row_no = len(self.df_economic_factors)
            self.logger.warning(
                "Row 11 not available after shift; using last available row (%s) for appended values.",
                source_row_no,
            )

        new_gdp = float(source_row['GDP']) if pd.notna(source_row['GDP']) else 0.0
        new_cpi = float(source_row['CPI']) if pd.notna(source_row['CPI']) else 0.0
        new_rf = float(source_row['Rf']) if pd.notna(source_row['Rf']) else 0.0
        new_unemp = float(source_row['Unemp']) if pd.notna(source_row['Unemp']) else 0.0

        new_row = pd.DataFrame({
            'Year':  [new_year],
            'GDP':   [new_gdp],
            'CPI':   [new_cpi],
            'Rf':    [new_rf],
            'Unemp': [new_unemp],
        })
        self.df_economic_factors = pd.concat(
            [self.df_economic_factors, new_row], ignore_index=True
        )
        self.logger.info(
            "Appended new row at end: Year=%s, GDP=%s, CPI=%s, Rf=%s, Unemp=%s (copied from row %s)",
            new_year,
            new_gdp,
            new_cpi,
            new_rf,
            new_unemp,
            source_row_no,
        )

        # Step 4: Update the row matching the user-entered year with user values
        if year in self.df_economic_factors['Year'].values:
            idx = self.df_economic_factors[self.df_economic_factors['Year'] == year].index[0]
            self.df_economic_factors.loc[idx] = [year, gdp, cpi, rf, unemp]
            self.logger.info(f"Updated Year={year} with: GDP={gdp}, CPI={cpi}, Rf={rf}, Unemp={unemp}")
        else:
            self.logger.warning(
                f"Year {year} not found in dataframe after shifting. "
                f"Available years: {list(self.df_economic_factors['Year'].values)}"
            )
    
    def update_without_shifting(self, year, gdp, cpi, rf, unemp):
        """
        Update dataframe without shifting (for March)
        Simply update or append the new data
        
        Args:
            year (int): Year value
            gdp (float): GDP value
            cpi (float): CPI value
            rf (float): Risk-free rate value
            unemp (float): Unemployment value
        """
        self.logger.info("=" * 80)
        self.logger.info("MARCH UPDATE - Updating Without Shifting")
        self.logger.info("=" * 80)
        
        if year in self.df_economic_factors['Year'].values:
            # Update existing year
            idx = self.df_economic_factors[self.df_economic_factors['Year'] == year].index[0]
            self.df_economic_factors.loc[idx] = [year, gdp, cpi, rf, unemp]
            self.logger.info(f"Updated existing row for year {year}")
        else:
            # Append new row
            new_row = pd.DataFrame({
                'Year': [year],
                'GDP': [gdp],
                'CPI': [cpi],
                'Rf': [rf],
                'Unemp': [unemp]
            })
            self.df_economic_factors = pd.concat([self.df_economic_factors, new_row], ignore_index=True)
            self.logger.info(f"Appended new row for year {year}")
        
        self.logger.info(f"New data: Year={year}, GDP={gdp}, CPI={cpi}, Rf={rf}, Unemp={unemp}")
    
    def update_economic_factors(self, running_date_str, year, gdp, cpi, rf, unemp):
        """
        Update economic factors based on running date
        
        Args:
            running_date_str (str): Running date in format DD.MM.YYYY
            year (int): Year value
            gdp (float): GDP value
            cpi (float): CPI value
            rf (float): Risk-free rate value
            unemp (float): Unemployment value
        """
        running_date = self.parse_running_date(running_date_str)
        month = running_date.month
        
        self.logger.info(f"Update parameters: Year={year}, GDP={gdp}, CPI={cpi}, Rf={rf}, Unemp={unemp}")
        
        if month == 9:  # September
            self.update_with_shifting(year, gdp, cpi, rf, unemp)
        elif month == 3:  # March
            self.update_without_shifting(year, gdp, cpi, rf, unemp)
        else:
            self.logger.warning(f"Month {month} is neither March (3) nor September (9)")
            self.logger.info("Applying default update (without shifting)")
            self.update_without_shifting(year, gdp, cpi, rf, unemp)
    
    def save_to_excel(self):
        """
        Save updated dataframe back into the original input file in-place.
        Uses openpyxl to update only the data cells (rows 3-15, cols A-E)
        so all other sheets, formatting, formulas, and styles are preserved.
        Percentage-formatted cells (GDP, CPI, Rf, Unemp) store decimal values
        (e.g., 0.065 is displayed as 6.5%%) — values are written as-is.
        """
        if not self.target_file or self.df_economic_factors is None:
            return
        
        self.logger.info("=" * 80)
        self.logger.info(f"Saving updated data back to original file: {self.target_file}")
        self.logger.info("=" * 80)
        
        try:
            # Load the full workbook preserving all formatting and styles
            wb = openpyxl.load_workbook(self.target_file, keep_links=False)
            ws = wb['Economic Factors']
            
            # Column positions in Excel (1-indexed)
            # A=Year, B=GDP, C=CPI, D=Rf, E=Unemp
            COL_MAP = {
                'Year':  1,
                'GDP':   2,
                'CPI':   3,
                'Rf':    4,
                'Unemp': 5,
            }
            START_ROW = 3  # Data begins at Excel row 3
            
            for df_idx, row in self.df_economic_factors.iterrows():
                excel_row = START_ROW + df_idx
                for col_name, col_num in COL_MAP.items():
                    raw_val = row[col_name]
                    if col_name == 'Year':
                        # Year stored as plain integer; NaN becomes 0
                        cell_value = int(raw_val) if pd.notna(raw_val) else 0
                    else:
                        # GDP, CPI, Rf, Unemp are percentage-formatted cells;
                        # Excel stores the decimal (0.065 = 6.5%); NaN becomes 0
                        cell_value = float(raw_val) if pd.notna(raw_val) else 0
                    ws.cell(row=excel_row, column=col_num).value = cell_value
            
            wb.save(self.target_file)
            self.logger.info(f"Successfully saved to: {self.target_file}")
            self.logger.info(
                f"Updated {len(self.df_economic_factors)} rows "
                f"(Excel rows {START_ROW} to {START_ROW + len(self.df_economic_factors) - 1})"
            )
        
        except PermissionError:
            self.logger.error(
                "Permission denied — the file is currently open in Excel. "
                "Please close it and run again."
            )
            raise
        except Exception as e:
            self.logger.error(f"Error saving to Excel: {str(e)}")
            raise
    
    def display_summary(self):
        """Display summary of the loaded data"""
        if self.df_economic_factors is not None:
            self.logger.info("\nAll Economic Factors Data:")
            self.logger.info("\n" + self.df_economic_factors.to_string())
    
    def run(self):
        """
        Main execution method
        
        Returns:
            pd.DataFrame: DataFrame containing economic factors data
        """
        self.logger.info("=" * 80)
        self.logger.info("Starting PD Loan Economic Factors Analysis")
        self.logger.info("=" * 80)
        
        # Find PD Loan file
        if not self.find_pd_loan_file():
            return None
        
        # Load Economic Factors
        self.load_economic_factors()
        
        # Display summary
        self.display_summary()
        
        self.logger.info("=" * 80)
        self.logger.info("Process completed successfully")
        self.logger.info("=" * 80)
        
        return self.df_economic_factors

    def run_step_2_historic_pd(self):
        """Step 2: Extract M row 5..10 values for all PD categories from Historic PD file."""
        self.logger.info("=" * 80)
        self.logger.info("Step 2 - Historic PD Category Extraction")
        self.logger.info("=" * 80)

        extractor = HistoricPDPivotExtractor(self.input_folder, self.logger)
        historic_file = extractor.find_historic_file("Historic PD Calculation")
        self.logger.info("Using Historic PD file: %s", historic_file)

        self.historic_pd_frames = extractor.extract_all_categories()
        self.logger.info("Step 2 completed. Extracted %s category dataframes.", len(self.historic_pd_frames))

        return self.historic_pd_frames

    def run_step_3_pd_category_update(self, year, running_date):
        """
        Step 3: Update PD category sheets with historic PD values.
        
        For September: Shift columns B:K leftward, add new year and M_values
        For March: Update M_values in place without shifting
        
        Args:
            year (int): Year value to update
            running_date (datetime): Parsed running date for determining month
        """
        if not self.historic_pd_frames:
            self.logger.warning("No historic PD frames available. Run Step 2 first.")
            return

        self.logger.info("=" * 80)
        self.logger.info("Step 3 - PD Category Sheet Updates")
        self.logger.info("=" * 80)

        import re
        
        month = running_date.month
        pd_loan_file = self.target_file
        
        if not pd_loan_file:
            self.logger.error("No PD Loan file identified. Run Step 1 first.")
            return

        try:
            wb = openpyxl.load_workbook(pd_loan_file)
            available_sheets = wb.sheetnames
            self.logger.info("Available sheets in workbook: %s", available_sheets)

            def normalize_sheet_name(name):
                return re.sub(r"\s+", "", str(name)).lower()

            def find_sheet_name_by_label(label):
                target = normalize_sheet_name(label)
                for s in available_sheets:
                    if normalize_sheet_name(s) == target:
                        return s
                return None

            def apply_shift_and_write_k_values(ws, year_value, k_values):
                # Step 1: Clear B2:B7
                for row in range(2, 8):
                    ws.cell(row=row, column=2).value = None

                # Step 2: Shift columns C:K leftward (C→B, D→C, ..., K→J)
                for row in range(2, 8):
                    for col in range(3, 12):  # Columns C(3) to K(11)
                        source_val = ws.cell(row=row, column=col).value
                        ws.cell(row=row, column=col - 1).value = source_val

                # Step 3: Clear column K rows 2-7
                for row in range(2, 8):
                    ws.cell(row=row, column=11).value = None

                # Step 4: Paste year to K2
                ws.cell(row=2, column=11).value = year_value

                # Step 5: Paste values to K3:K7
                for idx, val in enumerate(k_values):
                    ws.cell(row=3 + idx, column=11).value = val

            def increment_year_cells_b_to_k(ws):
                """
                Increment year-like values by 1 for:
                Row 9, 14, 19, 24 and Columns B to K.
                """
                updates = 0
                for target_row in [9, 14, 19, 24]:
                    for col in range(2, 12):  # B(2) to K(11)
                        raw_val = ws.cell(row=target_row, column=col).value
                        try:
                            year_val = int(float(raw_val))
                        except (TypeError, ValueError):
                            continue

                        if 1900 <= year_val <= 2200:
                            ws.cell(row=target_row, column=col).value = year_val + 1
                            updates += 1
                return updates

            def set_k_rows_to_year(ws, year_value):
                """Set K9, K14, K19, K24 to the running year."""
                updates = 0
                for target_row in [9, 14, 19, 24]:
                    ws.cell(row=target_row, column=11).value = year_value
                    updates += 1
                return updates

            for category, df_category in self.historic_pd_frames.items():
                # Build regex pattern to match sheet name with flexible spacing
                # Pattern: "PD" + optional spaces + "-" + optional spaces + category
                pattern = r'PD\s*-\s*' + re.escape(category)
                
                # Find the matching sheet name
                sheet_name = None
                for available_sheet in available_sheets:
                    if re.match(pattern, available_sheet, re.IGNORECASE):
                        sheet_name = available_sheet
                        break
                
                if not sheet_name:
                    self.logger.warning("Sheet matching 'PD*-*%s' not found. Skipping.", category)
                    continue

                ws = wb[sheet_name]
                self.logger.info("Updating sheet: %s (matched category: %s)", sheet_name, category)

                # Extract M_values for DC_BUCKET 1-5 (excel_row 5-9)
                m_values = []
                for idx in range(5):  # Rows 0-4 in dataframe (DC_BUCKET 1-5)
                    if idx < len(df_category):
                        m_val = df_category.iloc[idx]['M_value']
                        m_values.append(m_val if pd.notna(m_val) else 0)
                    else:
                        m_values.append(0)

                if month == 9:  # September - Shift method
                    self.logger.info("  Using SHIFT method (September)")
                    apply_shift_and_write_k_values(ws, year, m_values)
                    self.logger.info("  Pasted M_values to K3:K7: %s", m_values)

                elif month == 3:  # March - Edit method
                    self.logger.info("  Using EDIT method (March)")
                    
                    # Just update M_values in K3:K7 without shifting
                    for idx, m_val in enumerate(m_values):
                        ws.cell(row=3 + idx, column=11).value = m_val
                    self.logger.info("  Updated M_values in K3:K7: %s", m_values)

                else:
                    self.logger.warning("  Month %s is neither March nor September. Applying edit method.", month)
                    # Default to edit method
                    for idx, m_val in enumerate(m_values):
                        ws.cell(row=3 + idx, column=11).value = m_val
                    self.logger.info("  Updated M_values in K3:K7: %s", m_values)

                # Only increment year cells in September
                if month == 9:
                    year_updates = increment_year_cells_b_to_k(ws)
                    self.logger.info("  Incremented year value(s) in B:K for rows 9,14,19,24: %s", year_updates)
                else:
                    self.logger.info("  Skipping year increment (only done in September)")

            # Additional fixed sheets update with 0 values
            fixed_sheet_labels = ["PD - Refinance", "PD - Factoring"]
            zero_values = [0, 0, 0, 0, 0]

            for fixed_label in fixed_sheet_labels:
                fixed_sheet_name = find_sheet_name_by_label(fixed_label)
                if not fixed_sheet_name:
                    self.logger.warning("Fixed sheet '%s' not found. Skipping.", fixed_label)
                    continue

                ws_fixed = wb[fixed_sheet_name]
                self.logger.info("Updating fixed sheet: %s", fixed_sheet_name)

                if month == 9:
                    self.logger.info("  Using SHIFT method (September) with zero values")
                    apply_shift_and_write_k_values(ws_fixed, year, zero_values)
                    self.logger.info("  Pasted zero values to K3:K7")
                elif month == 3:
                    self.logger.info("  Using EDIT method (March) with zero values")
                    for idx, val in enumerate(zero_values):
                        ws_fixed.cell(row=3 + idx, column=11).value = val
                    self.logger.info("  Updated zero values in K3:K7")
                else:
                    self.logger.warning("  Month %s is neither March nor September. Applying edit method with zero values.", month)
                    for idx, val in enumerate(zero_values):
                        ws_fixed.cell(row=3 + idx, column=11).value = val
                    self.logger.info("  Updated zero values in K3:K7")

                # Only increment year cells in September
                if month == 9:
                    fixed_year_updates = increment_year_cells_b_to_k(ws_fixed)
                    self.logger.info(
                        "  Incremented year value(s) in B:K for rows 9,14,19,24 (fixed sheet): %s",
                        fixed_year_updates,
                    )

                    if normalize_sheet_name(fixed_label) == normalize_sheet_name("PD - Factoring"):
                        k_year_updates = set_k_rows_to_year(ws_fixed, year)
                        self.logger.info(
                            "  Set K9,K14,K19,K24 to running year %s for PD - Factoring (updated cells: %s)",
                            year,
                            k_year_updates,
                        )
                else:
                    self.logger.info("  Skipping year increment for fixed sheet (only done in September)")

            # Save the workbook
            wb.save(pd_loan_file)
            self.logger.info("Successfully saved updated PD category sheets to: %s", pd_loan_file)
            self.logger.info("Step 3 completed. Updated %s category sheets.", len(self.historic_pd_frames))

        except PermissionError:
            self.logger.error(
                "Permission denied — the file is currently open in Excel. "
                "Please close it and run again."
            )
            raise
        except Exception as e:
            self.logger.error(f"Error updating PD category sheets: {str(e)}")
            raise

    def run_step_4_increment_economic_factor_years(self):
        """
        Step 4: Find Economic Factor(s) sheet in PD Loan file and increment
        year values in row 3 from column C onward (e.g., 2024 -> 2025).
        """
        self.logger.info("=" * 80)
        self.logger.info("Step 4 - Increment Economic Factor Years")
        self.logger.info("=" * 80)

        if not self.target_file:
            self.logger.error("No PD Loan file identified. Run Step 1 first.")
            return

        try:
            wb = openpyxl.load_workbook(self.target_file)

            # Use exact sheet name as requested
            if "Economic Factor" not in wb.sheetnames:
                self.logger.warning(
                    "Sheet 'Economic Factor' not found. Available sheets: %s",
                    wb.sheetnames,
                )
                return

            ws = wb["Economic Factor"]
            self.logger.info("Using sheet: Economic Factor")

            target_rows = [3, 12, 21]
            total_updates = 0

            for target_row in target_rows:
                raw_rows = []
                for col in range(3, 14):  # C(3) to M(13)
                    raw_val = ws.cell(row=target_row, column=col).value
                    raw_rows.append(
                        {
                            "row": target_row,
                            "column": openpyxl.utils.get_column_letter(col),
                            "column_number": col,
                            "raw_value": raw_val,
                        }
                    )

                df_row_raw = pd.DataFrame(raw_rows)
                df_row_non_empty = df_row_raw[df_row_raw["raw_value"].notna()].copy()

                self.logger.info("Step 4 row %s raw dataframe shape: %s", target_row, df_row_raw.shape)
                if not df_row_non_empty.empty:
                    self.logger.info(
                        "Step 4 row %s non-empty values from columns C to M:\n%s",
                        target_row,
                        df_row_non_empty.to_string(index=False),
                    )
                else:
                    self.logger.warning(
                        "Step 4 test dataframe has no non-empty values in row %s for columns C to M.",
                        target_row,
                    )

                row_updates = 0
                for _, row_data in df_row_non_empty.iterrows():
                    raw_val = row_data["raw_value"]
                    col_num = int(row_data["column_number"])

                    try:
                        year_val = int(float(raw_val))
                    except (TypeError, ValueError):
                        continue

                    if 1900 <= year_val <= 2200:
                        ws.cell(row=target_row, column=col_num).value = year_val + 1
                        row_updates += 1

                total_updates += row_updates
                self.logger.info("Step 4 row %s incremented year value(s): %s", target_row, row_updates)

            if total_updates > 0:
                wb.save(self.target_file)
                self.logger.info(
                    "Step 4 completed. Incremented total %s year value(s) in rows 3, 12, and 21 for columns C to M.",
                    total_updates,
                )
            else:
                self.logger.warning("No year values found to increment in rows 3, 12, and 21 for columns C to M.")

        except PermissionError:
            self.logger.error(
                "Permission denied — the file is currently open in Excel. "
                "Please close it and run again."
            )
            raise
        except Exception as e:
            self.logger.error(f"Error in Step 4 year increment: {str(e)}")
            raise

    def run_step_5_paste_pd_weighted_data(self, year, running_date):
        """
        Step 5: Paste PD weighted data with shift logic.

        Rules:
        - PD - RL:  AB->AK range, year in AK1, RL excel_row=10 in AK2
        - PD - SME: AA->AJ range, year in AJ1, SME excel_row=10 in AJ2
        - PD -TL:  AA->AJ range, year in AJ1, SME excel_row=10 in AJ2
        - PD - Refinance: AB->AK range, year in AK1, fixed 0.0 in AK2
        """
        self.logger.info("=" * 80)
        self.logger.info("Step 5 - Paste PD Weighted Data")
        self.logger.info("=" * 80)

        if not self.target_file:
            self.logger.error("No PD Loan file identified. Run Step 1 first.")
            return

        if not self.historic_pd_frames:
            self.logger.error("No historic PD frames available. Run Step 2 first.")
            return

        month = running_date.month

        try:
            wb = openpyxl.load_workbook(self.target_file)
            available_sheets = wb.sheetnames

            def normalize_sheet_name(name):
                return "".join(str(name).split()).lower()

            def find_sheet_name_by_label(label):
                target = normalize_sheet_name(label)
                for s in available_sheets:
                    if normalize_sheet_name(s) == target:
                        return s
                return None

            def get_weighted_value(category_name):
                df_cat = self.historic_pd_frames.get(category_name)
                if df_cat is None:
                    self.logger.warning("Category '%s' not found in historic frames. Using 0.", category_name)
                    return 0

                row_10 = df_cat[df_cat["excel_row"] == 10]
                if row_10.empty:
                    self.logger.warning("Category '%s' has no excel_row=10. Using 0.", category_name)
                    return 0

                val = row_10.iloc[0]["M_value"]
                return 0 if pd.isna(val) else float(val)

            def apply_shift_and_paste(ws, start_col_letter, end_col_letter, year_value, weighted_value):
                start_col = openpyxl.utils.column_index_from_string(start_col_letter)
                end_col = openpyxl.utils.column_index_from_string(end_col_letter)

                # Clear start column rows 1..2
                for r in range(1, 3):
                    ws.cell(row=r, column=start_col).value = None

                # Shift one column left inside the range for rows 1..2
                for r in range(1, 3):
                    for col in range(start_col + 1, end_col + 1):
                        ws.cell(row=r, column=col - 1).value = ws.cell(row=r, column=col).value

                # Clear end column rows 1..2
                for r in range(1, 3):
                    ws.cell(row=r, column=end_col).value = None

                # Paste year and weighted value
                ws.cell(row=1, column=end_col).value = year_value
                ws.cell(row=2, column=end_col).value = weighted_value

            def paste_without_shift_by_year(ws, start_col_letter, end_col_letter, year_value, weighted_value):
                start_col = openpyxl.utils.column_index_from_string(start_col_letter)
                end_col = openpyxl.utils.column_index_from_string(end_col_letter)

                for col in range(start_col, end_col + 1):
                    header_val = ws.cell(row=1, column=col).value
                    try:
                        header_year = int(float(header_val))
                    except (TypeError, ValueError):
                        continue

                    if header_year == year_value:
                        ws.cell(row=2, column=col).value = weighted_value
                        return openpyxl.utils.get_column_letter(col)

                return None

            step5_rules = [
                {
                    "sheet_label": "PD - RL",
                    "start_col": "AB",
                    "end_col": "AK",
                    "category": "RL",
                },
                {
                    "sheet_label": "PD - SME",
                    "start_col": "AA",
                    "end_col": "AJ",
                    "category": "SME",
                },
                {
                    "sheet_label": "PD -TL",
                    "start_col": "AA",
                    "end_col": "AJ",
                    "category": "TL",
                },
                {
                    "sheet_label": "PD - Refinance",
                    "start_col": "AB",
                    "end_col": "AK",
                    "fixed_weighted_value": 0.0,
                },
            ]

            for rule in step5_rules:
                matched_sheet = find_sheet_name_by_label(rule["sheet_label"])
                if not matched_sheet:
                    self.logger.warning("Step 5 sheet '%s' not found. Skipping.", rule["sheet_label"])
                    continue

                ws = wb[matched_sheet]
                if "fixed_weighted_value" in rule:
                    weighted_val = float(rule["fixed_weighted_value"])
                else:
                    weighted_val = get_weighted_value(rule["category"])

                if month == 9:
                    # September: shift and append latest year/value at right edge
                    apply_shift_and_paste(
                        ws,
                        rule["start_col"],
                        rule["end_col"],
                        year,
                        weighted_val,
                    )
                    self.logger.info(
                        "Step 5 (September) updated sheet '%s': %s1=%s, %s2=%s",
                        matched_sheet,
                        rule["end_col"],
                        year,
                        rule["end_col"],
                        weighted_val,
                    )
                elif month == 3:
                    # March: no shift; update row-2 value under the matching year header
                    updated_col = paste_without_shift_by_year(
                        ws,
                        rule["start_col"],
                        rule["end_col"],
                        year,
                        weighted_val,
                    )
                    if updated_col:
                        self.logger.info(
                            "Step 5 (March) updated sheet '%s': %s2=%s for year %s",
                            matched_sheet,
                            updated_col,
                            weighted_val,
                            year,
                        )
                    else:
                        self.logger.warning(
                            "Step 5 (March) year %s not found in row 1 range %s:%s for sheet '%s'.",
                            year,
                            rule["start_col"],
                            rule["end_col"],
                            matched_sheet,
                        )
                else:
                    self.logger.warning(
                        "Step 5 month %s is neither March nor September; no update applied for sheet '%s'.",
                        month,
                        matched_sheet,
                    )

            wb.save(self.target_file)
            self.logger.info("Step 5 completed successfully.")

        except PermissionError:
            self.logger.error(
                "Permission denied — the file is currently open in Excel. "
                "Please close it and run again."
            )
            raise
        except Exception as e:
            self.logger.error(f"Error in Step 5 weighted data update: {str(e)}")
            raise

    def run_step_6_final_save_and_rename(self, input_year, running_date):
        """
        Final step:
        - September: rename PD Loan file using (input_year + 1)
        - Other months: keep filename unchanged (save only)
        """
        self.logger.info("=" * 80)
        self.logger.info("Step 6 - Final Save and Rename")
        self.logger.info("=" * 80)

        if not self.target_file:
            self.logger.error("No PD Loan file identified. Cannot run final rename step.")
            return None

        month = running_date.month
        if month != 9:
            self.logger.info("Non-September run: file kept as-is (save only): %s", self.target_file)
            return self.target_file

        import re

        current_path = self.target_file
        folder = os.path.dirname(current_path)
        file_name = os.path.basename(current_path)
        next_year = int(input_year) + 1

        # Primary replacement: entered year token -> entered year + 1
        year_token_pattern = rf"(?<!\d){re.escape(str(input_year))}(?!\d)"
        new_file_name = re.sub(year_token_pattern, str(next_year), file_name, count=1)

        # Fallback replacement: first 4-digit year token
        if new_file_name == file_name:
            new_file_name = re.sub(r"(?<!\d)(?:19|20)\d{2}(?!\d)", str(next_year), file_name, count=1)

        if new_file_name == file_name:
            self.logger.warning(
                "Could not find a year token to rename in filename: %s. Keeping original name.",
                file_name,
            )
            return self.target_file

        new_path = os.path.join(folder, new_file_name)

        if os.path.abspath(new_path) == os.path.abspath(current_path):
            self.logger.info("Filename unchanged after rename logic: %s", current_path)
            return self.target_file

        if os.path.exists(new_path):
            self.logger.warning(
                "Target rename file already exists: %s. Keeping current file name: %s",
                new_path,
                current_path,
            )
            return self.target_file

        os.rename(current_path, new_path)
        self.target_file = new_path
        self.logger.info("Renamed file for September run: %s -> %s", current_path, new_path)
        return new_path

    def paste_scenarios_to_scorecard(self, normal_scenario, best_case_scenario, worst_case_scenario):
        """
        Paste scenario values to the "Score card - RL" sheet
        B53: Normal Scenario
        B54: Best case Scenario
        B55: Worst case Scenario
        
        Args:
            normal_scenario (float): Normal Scenario value
            best_case_scenario (float): Best case Scenario value
            worst_case_scenario (float): Worst case Scenario value
        """
        self.logger.info("=" * 80)
        self.logger.info("Pasting Scenario Values to Score card - RL Sheet")
        self.logger.info("=" * 80)
        
        if not self.target_file:
            self.logger.error("No PD Loan file identified. Cannot paste scenario values.")
            return False
        
        try:
            # Load the workbook
            wb = openpyxl.load_workbook(self.target_file, keep_links=False)
            
            # Check if "Score card - RL" sheet exists
            if "Score card - RL" not in wb.sheetnames:
                self.logger.error("Sheet 'Score card - RL' not found in workbook. Available sheets: %s", wb.sheetnames)
                wb.close()
                return False
            
            ws = wb["Score card - RL"]
            
            # Paste values to cells B53, B54, B55 (column 2)
            ws.cell(row=53, column=2).value = normal_scenario
            ws.cell(row=54, column=2).value = best_case_scenario
            ws.cell(row=55, column=2).value = worst_case_scenario
            
            self.logger.info("Pasted Normal Scenario: %s to B53", normal_scenario)
            self.logger.info("Pasted Best case Scenario: %s to B54", best_case_scenario)
            self.logger.info("Pasted Worst case Scenario: %s to B55", worst_case_scenario)
            
            # Save the workbook
            wb.save(self.target_file)
            wb.close()
            
            self.logger.info("Successfully saved scenario values to: %s", self.target_file)
            return True
            
        except Exception as e:
            self.logger.error("Error pasting scenario values: %s", str(e))
            return False


def main():
    """Main entry point"""
    # Setup argument parser
    parser = argparse.ArgumentParser(
        description='PD Loan Economic Factors Analysis',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # View current data
  python PD_Loan.py
  
  # Update with March date (no shifting)
  python PD_Loan.py --running_date 15.03.2025 --year 2025 --gdp 0.045 --cpi 0.055 --rf 0.0850 --unemp 0.0480
  
  # Update with September date (with shifting)
  python PD_Loan.py --running_date 10.09.2025 --year 2025 --gdp 0.048 --cpi 0.060 --rf 0.0900 --unemp 0.0450
        """
    )
    
    parser.add_argument('--running_date', type=str, help='Running date in format DD.MM.YYYY (e.g., 9.03.2025)')
    parser.add_argument('--year', type=int, help='Year value to update (optional; defaults to year in --running_date)')
    parser.add_argument('--gdp', type=float, help='GDP value')
    parser.add_argument('--cpi', type=float, help='CPI value')
    parser.add_argument('--rf', type=float, help='Risk-free rate value')
    parser.add_argument('--unemp', type=float, help='Unemployment value')
    parser.add_argument('--normal_scenario', type=float, help='Normal Scenario value for Score card - RL sheet (B53)')
    parser.add_argument('--best_case_scenario', type=float, help='Best case Scenario value for Score card - RL sheet (B54)')
    parser.add_argument('--worst_case_scenario', type=float, help='Worst case Scenario value for Score card - RL sheet (B55)')
    
    args = parser.parse_args()
    
    input_folder = r"C:\MY\Overdime\IFRS\IFRS-main\Input Files\PD"
    log_folder = r"C:\MY\Overdime\IFRS\IFRS-main\Scripts\Logs"
    
    # Create analyzer instance and run
    analyzer = PDLoanAnalyzer(input_folder, log_folder)
    df = analyzer.run()
    
    # Check if update parameters are provided
    if args.running_date:
        # Validate required economic parameters (year can be derived from running_date)
        if not all([args.gdp is not None, args.cpi is not None, 
                   args.rf is not None, args.unemp is not None]):
            analyzer.logger.error("Parameters (--gdp, --cpi, --rf, --unemp) are required when --running_date is specified")
            return None
        
        # Parse running date for Step 3
        running_date = analyzer.parse_running_date(args.running_date)
        effective_year = args.year if args.year is not None else running_date.year
        if args.year is None:
            analyzer.logger.info("--year not provided. Using year from --running_date: %s", effective_year)
        
        # Update economic factors
        analyzer.update_economic_factors(
            args.running_date,
            effective_year,
            args.gdp,
            args.cpi,
            args.rf,
            args.unemp
        )
        
        # Display updated data
        analyzer.logger.info("\nUpdated Economic Factors Data:")
        analyzer.logger.info("\n" + analyzer.df_economic_factors.to_string())
        
        # Save to Excel
        analyzer.save_to_excel()
        
        df = analyzer.df_economic_factors

    # Step 2: Historic PD extraction from 03.PD_Pivot logic (M row 5..10)
    analyzer.run_step_2_historic_pd()
    
    # Step 3: Update PD category sheets with historic PD values
    if args.running_date:
        analyzer.run_step_3_pd_category_update(effective_year, running_date)
        if running_date.month == 9:
            analyzer.run_step_4_increment_economic_factor_years()
        else:
            analyzer.logger.info("Step 4 skipped: it runs only for September month.")
        analyzer.run_step_5_paste_pd_weighted_data(effective_year, running_date)
        analyzer.run_step_6_final_save_and_rename(effective_year, running_date)
        
        # Step 7: Paste scenario values to Score card - RL sheet if provided
        if args.normal_scenario is not None or args.best_case_scenario is not None or args.worst_case_scenario is not None:
            analyzer.paste_scenarios_to_scorecard(
                args.normal_scenario if args.normal_scenario is not None else 0,
                args.best_case_scenario if args.best_case_scenario is not None else 0,
                args.worst_case_scenario if args.worst_case_scenario is not None else 0
            )
    
    return df


if __name__ == "__main__":
    df = main()