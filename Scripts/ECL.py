import glob
import os
import logging
from datetime import datetime
from typing import List, Optional, Tuple

import openpyxl
import pandas as pd


INPUT_PD_FOLDER = r"C:\MY\Overdime\IFRS\IFRS-main\Input Files\PD"
LOG_FOLDER = r"C:\MY\Overdime\IFRS\IFRS-main\Scripts\Logs"
TARGET_SHEET = "PD weighted  year"
START_COL = 2   # B
END_COL = 37    # AK
START_ROW = 36
END_ROW = 41


def setup_logger() -> logging.Logger:
	os.makedirs(LOG_FOLDER, exist_ok=True)
	log_file = os.path.join(LOG_FOLDER, f"ECL_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
	logger = logging.getLogger("ECL")
	logger.setLevel(logging.INFO)
	logger.handlers.clear()

	formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
	file_handler = logging.FileHandler(log_file)
	file_handler.setFormatter(formatter)
	stream_handler = logging.StreamHandler()
	stream_handler.setFormatter(formatter)

	logger.addHandler(file_handler)
	logger.addHandler(stream_handler)
	logger.info("Log file: %s", log_file)
	return logger


def find_pd_loan_file(input_folder: str) -> Optional[str]:
	pattern = os.path.join(input_folder, "**", "*PD Loan*.xlsx*")
	matches = [
		path for path in glob.glob(pattern, recursive=True)
		if not os.path.basename(path).startswith("~$")
	]
	if not matches:
		return None

	# Use latest modified PD Loan file.
	matches.sort(key=os.path.getmtime, reverse=True)
	return matches[0]


def find_pd_lease_file(input_folder: str) -> Optional[str]:
	pattern = os.path.join(input_folder, "**", "*PD Lease*.xlsx*")
	matches = [
		path for path in glob.glob(pattern, recursive=True)
		if not os.path.basename(path).startswith("~$")
	]
	if not matches:
		return None

	# Use latest modified PD Lease file.
	matches.sort(key=os.path.getmtime, reverse=True)
	return matches[0]


def read_pd_weighted_year_block_with_range(
	workbook_file: str,
	start_col: int,
	end_col: int,
	start_row: int,
	end_row: int,
) -> Tuple[pd.DataFrame, List[List[str]]]:
	wb = openpyxl.load_workbook(workbook_file, data_only=True)
	if TARGET_SHEET not in wb.sheetnames:
		raise ValueError(f"Sheet '{TARGET_SHEET}' not found in {workbook_file}")

	ws = wb[TARGET_SHEET]
	data_rows = []
	number_formats = []
	for row in range(start_row, end_row + 1):
		row_values = [
			ws.cell(row=row, column=col).value
			for col in range(start_col, end_col + 1)
		]
		row_formats = [
			ws.cell(row=row, column=col).number_format
			for col in range(start_col, end_col + 1)
		]
		data_rows.append(row_values)
		number_formats.append(row_formats)

	columns = [openpyxl.utils.get_column_letter(col) for col in range(start_col, end_col + 1)]
	index = [f"row{r}" for r in range(start_row, end_row + 1)]
	return pd.DataFrame(data_rows, columns=columns, index=index), number_formats


def read_pd_weighted_year_block(pd_loan_file: str) -> Tuple[pd.DataFrame, List[List[str]]]:
	return read_pd_weighted_year_block_with_range(
		pd_loan_file,
		START_COL,
		END_COL,
		START_ROW,
		END_ROW,
	)


def find_ecl_model_file(input_folder: str, keyword: str = "ECL Model") -> Optional[str]:
	pattern_xlsb = os.path.join(input_folder, "**", f"*{keyword}*.xlsb")
	matches_xlsb = [
		path for path in glob.glob(pattern_xlsb, recursive=True)
		if not os.path.basename(path).startswith("~$")
	]
	if not matches_xlsb:
		return None

	# Prefer the explicitly named file if it exists.
	preferred_name = "ECL Model_2025-08-18.xlsb"
	for path in matches_xlsb:
		if os.path.basename(path).strip().lower() == preferred_name.lower():
			return path

	# Prefer latest modified .xlsb ECL Model file.
	matches_xlsb.sort(key=os.path.getmtime, reverse=True)
	return matches_xlsb[0]


def clear_and_paste_to_ecl_model_range(df: pd.DataFrame, ecl_model_file: str, target_range: str) -> None:
	# Expected shape: 6 rows and same number of columns as the target range width.
	if df.shape[0] != 6:
		raise ValueError(f"Unexpected DataFrame rows {df.shape[0]}. Expected 6")

	# Convert to values-only matrix.
	values_matrix = []
	for _, row in df.iterrows():
		out_row = []
		for val in row.tolist():
			if pd.isna(val):
				out_row.append(0)
			else:
				out_row.append(val)
		values_matrix.append(out_row)

	# If the first pasted row has 0 in a column,
	# clear that whole column for all pasted rows.
	cols_to_clear = [idx for idx, val in enumerate(values_matrix[0]) if val == 0]
	for r_idx in range(len(values_matrix)):
		for c_idx in cols_to_clear:
			values_matrix[r_idx][c_idx] = None

	try:
		import win32com.client as win32
	except ImportError as e:
		raise ImportError("win32com.client is required to write to .xlsb files") from e

	excel = None
	wb = None
	try:
		excel = win32.Dispatch("Excel.Application")
		excel.Visible = False
		excel.DisplayAlerts = False
		excel.ScreenUpdating = False

		wb = excel.Workbooks.Open(os.path.abspath(ecl_model_file))

		target_sheet_name = "PD weighted  year"
		ws = None
		for i in range(1, wb.Sheets.Count + 1):
			if wb.Sheets(i).Name.strip() == target_sheet_name.strip():
				ws = wb.Sheets(i)
				break

		if ws is None:
			available = [wb.Sheets(i).Name for i in range(1, wb.Sheets.Count + 1)]
			raise ValueError(
				f"Sheet '{target_sheet_name}' not found in ECL Model. Available sheets: {available}"
			)

		# Clear target range and paste values.
		ws.Range(target_range).ClearContents()
		ws.Range(target_range).Value2 = tuple(tuple(row) for row in values_matrix)

		wb.Save()
	finally:
		try:
			if wb is not None:
				wb.Close(False)
		except Exception:
			pass
		try:
			if excel is not None:
				excel.ScreenUpdating = True
				excel.Quit()
		except Exception:
			pass


def clear_and_paste_to_ecl_model(df: pd.DataFrame, ecl_model_file: str) -> None:
	# Existing PD Loan mapping
	clear_and_paste_to_ecl_model_range(df, ecl_model_file, target_range="B13:AK18")


def main() -> None:
	logger = setup_logger()
	pd_loan_file = find_pd_loan_file(INPUT_PD_FOLDER)
	if not pd_loan_file:
		raise FileNotFoundError(f"No PD Loan workbook found in: {INPUT_PD_FOLDER}")

	df, number_formats = read_pd_weighted_year_block(pd_loan_file)
	ecl_model_file = find_ecl_model_file(INPUT_PD_FOLDER, keyword="ECL Model")
	if not ecl_model_file:
		raise FileNotFoundError(f"No ECL Model .xlsb workbook found in: {INPUT_PD_FOLDER}")

	clear_and_paste_to_ecl_model(df, ecl_model_file)

	# Next step: PD Lease -> ECL Model
	pd_lease_file = find_pd_lease_file(INPUT_PD_FOLDER)
	if not pd_lease_file:
		raise FileNotFoundError(f"No PD Lease workbook found in: {INPUT_PD_FOLDER}")

	# Lease range: B36:BM41
	lease_start_col = 2   # B
	lease_end_col = 65    # BM
	lease_df, lease_number_formats = read_pd_weighted_year_block_with_range(
		pd_lease_file,
		lease_start_col,
		lease_end_col,
		START_ROW,
		END_ROW,
	)

	# Paste Lease values into ECL hidden sheet range B3:BM8
	clear_and_paste_to_ecl_model_range(lease_df, ecl_model_file, target_range="B3:BM8")
	logger.info("PD Loan file: %s", pd_loan_file)
	logger.info("DataFrame from sheet 'PD weighted  year', range B36:AK41")
	logger.info("\n%s", df.to_string())
	logger.info("Pasted values to ECL Model file: %s", ecl_model_file)
	logger.info("Updated hidden sheet 'PD weighted  year' range: B13:AK18")
	logger.info("PD Lease file: %s", pd_lease_file)
	logger.info("DataFrame from sheet 'PD weighted  year', range B36:BM41")
	logger.info("\n%s", lease_df.to_string())
	logger.info("Pasted Lease values to ECL Model file: %s", ecl_model_file)
	logger.info("Updated hidden sheet 'PD weighted  year' range: B3:BM8")


if __name__ == "__main__":
	main()
